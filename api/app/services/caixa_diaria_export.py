"""Exportació de la caixa diària mensual a Excel i PDF, mateix format que la
plantilla d'Excel que portava la botiga abans (veure schemas.CaixaDiariaMesOut)."""

import io

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schemas import CaixaDiariaMesOut

COLUMNES = (
    ("date", "DIA"),
    ("card_21", "TARGETA 21%"),
    ("card_4", "TARGETA 4%"),
    ("cash_21", "EFECTIU 21%"),
    ("cash_4", "EFECTIU 4%"),
    ("bizum_21", "BIZUM 21%"),
    ("bizum_4", "BIZUM 4%"),
    ("paypal_21", "PAYPAL 21%"),
    ("paypal_4", "PAYPAL 4%"),
    ("transfer_21", "TRANSFER 21%"),
    ("cultural_voucher", "BONO CULTURAL"),
    ("total_dia", "TOTAL DIA"),
)


def generate_caixa_diaria_excel(mes: CaixaDiariaMesOut) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{mes.year}-{mes.mes:02d}"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    totals_fill = PatternFill("solid", fgColor="00B0F0")
    totals_font = Font(bold=True)

    for col, (_, label) in enumerate(COLUMNES, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, dia in enumerate(mes.dies, start=2):
        ws.cell(row=i, column=1, value=dia.date.strftime("%d-%m-%Y"))
        for col, (camp, _) in enumerate(COLUMNES[1:], start=2):
            valor = float(dia.total_dia) if camp == "total_dia" else float(getattr(dia, camp))
            ws.cell(row=i, column=col, value=valor).number_format = "#,##0.00"

    fila_totals = len(mes.dies) + 2
    ws.cell(row=fila_totals, column=1, value="TOTAL")
    for col, (camp, _) in enumerate(COLUMNES[1:], start=2):
        valor = float(mes.total_mes) if camp == "total_dia" else float(getattr(mes.totals, camp))
        cell = ws.cell(row=fila_totals, column=col, value=valor)
        cell.number_format = "#,##0.00"
        cell.fill = totals_fill
        cell.font = totals_font
    ws.cell(row=fila_totals, column=1).fill = totals_fill
    ws.cell(row=fila_totals, column=1).font = totals_font

    ws.column_dimensions[get_column_letter(1)].width = 12
    for col in range(2, len(COLUMNES) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_caixa_diaria_pdf(mes: CaixaDiariaMesOut) -> bytes:
    pdf = FPDF(format="A4", orientation="L")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, f"Caixa diaria - {mes.mes:02d}/{mes.year}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    col_width = 23
    dia_width = 20
    row_h = 5.5

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(dia_width, row_h, "DIA", border=1, align="C", fill=True)
    for _, label in COLUMNES[1:]:
        pdf.cell(col_width, row_h, label, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(0, 0, 0)
    for dia in mes.dies:
        pdf.cell(dia_width, row_h, dia.date.strftime("%d-%m-%Y"), border=1, align="C")
        for camp, _ in COLUMNES[1:-1]:
            pdf.cell(col_width, row_h, f"{getattr(dia, camp):.2f}", border=1, align="R")
        pdf.cell(col_width, row_h, f"{dia.total_dia:.2f}", border=1, align="R")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(0, 176, 240)
    pdf.cell(dia_width, row_h, "TOTAL", border=1, align="C", fill=True)
    for camp, _ in COLUMNES[1:-1]:
        pdf.cell(col_width, row_h, f"{getattr(mes.totals, camp):.2f}", border=1, align="R", fill=True)
    pdf.cell(col_width, row_h, f"{mes.total_mes:.2f}", border=1, align="R", fill=True)
    pdf.ln()

    return bytes(pdf.output())
